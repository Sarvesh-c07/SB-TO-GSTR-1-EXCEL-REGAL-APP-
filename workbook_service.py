from __future__ import annotations

import json
import os
import re
import shutil
import tempfile
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    ("particulars", "Particulars"), ("bill_no", "Bill No."), ("invoice_date", "Date"),
    ("sb_no", "SB NO."), ("sb_date", "SB DATE"), ("port_code", "Port Code"),
    ("foreign_amount", "$ / E"), ("exchange_rate", "Exch Rate"), ("inr", "INR"),
    ("taxable_value", "Value@5%"), ("igst", "IGST@5%"), ("fob", "FOB"),
    ("drawback", "DBK"), ("rodtep", "RoDTEP"),
]

ALIASES = {
    "particulars": {"particulars", "buyer", "consignee", "customer"},
    "bill_no": {"billno", "billnumber", "invoiceno", "invoicenumber"},
    "invoice_date": {"date", "invoicedate", "billdate"},
    "sb_no": {"sbno", "shippingbillno", "shippingbillnumber"},
    "sb_date": {"sbdate", "shippingbilldate"},
    "port_code": {"portcode", "port"},
    "foreign_amount": {"e", "foreignamount", "invoicevalue", "fcvalue", "forexamount"},
    "exchange_rate": {"exchrate", "exchangerate", "forexrate"},
    "inr": {"inr", "inrvalue", "rupeevalue"},
    "taxable_value": {"value5", "valueat5", "igstvalue", "taxablevalue"},
    "igst": {"igst5", "igstat5", "igstamount", "igst"},
    "fob": {"fob", "fobvalue"},
    "drawback": {"dbk", "drawback", "drawbackamount"},
    "rodtep": {"rodtep", "rodtepamount"},
}


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _field_for_header(value: Any) -> str | None:
    normalized = _norm(value)
    for field, aliases in ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _sheet_and_headers(workbook: Any):
    best = None
    for ws in workbook.worksheets:
        for row in range(1, min(ws.max_row, 60) + 1):
            mapping = {}
            for col in range(1, ws.max_column + 1):
                field = _field_for_header(ws.cell(row, col).value)
                if field and field not in mapping:
                    mapping[field] = col
            score = len(mapping)
            if best is None or score > best[0]:
                best = (score, ws, row, mapping)
    if not best or best[0] < 5:
        raise ValueError("Could not find the master table headers. Keep at least five recognized headers in one row.")
    return best[1], best[2], best[3]


def _total_row(ws: Any, header_row: int) -> int | None:
    for row in range(header_row + 1, ws.max_row + 1):
        values = [_norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "total" in values:
            return row
    return None


def _key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().upper()


def _as_date(value: Any) -> Any:
    if not value:
        return None
    if isinstance(value, (date, datetime)):
        return value
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return value


def _copy_style(source: Any, target: Any) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _extend_formula(formula: str, old_end: int, new_end: int) -> str:
    pattern = re.compile(rf"(\$?[A-Z]{{1,3}}\$?\d+:\$?[A-Z]{{1,3}}\$?){old_end}(?!\d)")
    return pattern.sub(rf"\g<1>{new_end}", formula)


def _prepare_new_row(ws: Any, row: int, previous: int, mapped_cols: set[int]) -> None:
    if previous < 1:
        return
    for col in range(1, ws.max_column + 1):
        source, target = ws.cell(previous, col), ws.cell(row, col)
        _copy_style(source, target)
        if col not in mapped_cols and target.value is None and isinstance(source.value, str) and source.value.startswith("="):
            try:
                target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
            except Exception:
                target.value = source.value


def create_new_master(path: Path, backup_dir: Path | None = None) -> None:
    if path.exists() and backup_dir is not None:
        backup_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, backup_dir / f"master_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "EXPORT"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.merge_cells("A1:N2")
    ws["A1"] = "3S PHARMACEUTICALS - EXPORT MASTER"
    ws["A1"].font = Font(name="Cambria", size=20, bold=True, color="F7E7B2")
    ws["A1"].fill = PatternFill("solid", fgColor="102A43")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"] = "Rule-based customs extraction | Editable living master"
    ws["A4"].font = Font(name="Calibri", italic=True, color="52606D")
    gold = "C9A227"
    thin = Side(style="thin", color="D6C9A5")
    for col, (_, label) in enumerate(HEADERS, start=1):
        cell = ws.cell(5, col, label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="174C45")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=gold))
    widths = [34, 12, 14, 15, 14, 13, 15, 13, 16, 16, 15, 16, 12, 13]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
        _copy_style(ws.cell(5, col), ws.cell(6, col))
        ws.cell(6, col).font = Font(name="Calibri", size=11, color="102A43")
        ws.cell(6, col).fill = PatternFill("solid", fgColor="FFFCF3")
        ws.cell(6, col).border = Border(bottom=thin)
    ws.row_dimensions[1].height = 31
    ws.row_dimensions[5].height = 34
    table = Table(displayName="ExportMaster", ref="A5:N6")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def master_status(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"ready": False}
    wb = load_workbook(path, data_only=False, read_only=False)
    ws, header_row, mapping = _sheet_and_headers(wb)
    key_col = mapping.get("bill_no")
    total = _total_row(ws, header_row) or ws.max_row + 1
    rows = sum(1 for r in range(header_row + 1, total) if key_col and ws.cell(r, key_col).value not in (None, ""))
    return {
        "ready": True, "filename": path.name, "sheet": ws.title, "header_row": header_row,
        "records": rows, "recognized_columns": len(mapping), "total_columns": ws.max_column,
    }


def append_rows(master_path: Path, rows: list[dict[str, Any]], audit_path: Path) -> dict[str, Any]:
    wb = load_workbook(master_path, data_only=False, keep_links=True)
    ws, header_row, mapping = _sheet_and_headers(wb)
    if "bill_no" not in mapping:
        raise ValueError("The master needs a Bill No. or Invoice No. column for safe duplicate handling.")
    key_col = mapping["bill_no"]
    mapped_cols = set(mapping.values())
    updates = additions = 0
    changed_rows = []

    for record in rows:
        total = _total_row(ws, header_row)
        data_limit = total if total else ws.max_row + 1
        existing = None
        for row in range(header_row + 1, data_limit):
            if _key(ws.cell(row, key_col).value) == _key(record.get("bill_no")) and _key(record.get("bill_no")):
                existing = row
                break
        if existing:
            target_row = existing
            updates += 1
            action = "updated"
        else:
            last_key_row = header_row
            for row in range(header_row + 1, data_limit):
                if ws.cell(row, key_col).value not in (None, ""):
                    last_key_row = row
            target_row = last_key_row + 1
            old_end = max(header_row, data_limit - 1)
            if total and target_row >= total:
                ws.insert_rows(total, 1)
                target_row = total
                for row in range(target_row + 1, min(ws.max_row, target_row + 8) + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row, col)
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            cell.value = _extend_formula(cell.value, old_end, old_end + 1)
            _prepare_new_row(ws, target_row, max(header_row + 1, target_row - 1), mapped_cols)
            additions += 1
            action = "added"

        for field, col in mapping.items():
            cell = ws.cell(target_row, col)
            if field == "inr":
                if "foreign_amount" in mapping and "exchange_rate" in mapping:
                    amount_col = get_column_letter(mapping["foreign_amount"])
                    rate_col = get_column_letter(mapping["exchange_rate"])
                    cell.value = f"={amount_col}{target_row}*{rate_col}{target_row}"
                    cell.number_format = '₹#,##0.00'
                continue
            if field not in record:
                continue
            value = record.get(field)
            if field in {"invoice_date", "sb_date"}:
                value = _as_date(value)
                cell.number_format = "d-mmm-yy"
            elif field == "foreign_amount":
                code = str(record.get("currency") or "USD").upper()
                symbol = {"USD": "$", "EUR": "€", "GBP": "£", "JPY": "¥"}.get(code, code)
                cell.number_format = f'"{symbol}"#,##0.00'
            elif field == "exchange_rate":
                cell.number_format = "0.0000"
            elif field in {"taxable_value", "igst", "fob"}:
                cell.number_format = '₹#,##0.00'
            elif field in {"drawback", "rodtep"}:
                cell.number_format = '₹#,##0'
            if value not in (None, ""):
                if field == "bill_no" and str(value).isdigit():
                    value = int(value)
                cell.value = value
        changed_rows.append({"row": target_row, "bill_no": record.get("bill_no"), "action": action})

    final_data_row = max([header_row] + [r for r in range(header_row + 1, ws.max_row + 1) if ws.cell(r, key_col).value not in (None, "")])
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row == header_row and min_col <= key_col <= max_col and final_data_row > max_row:
            table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{final_data_row}"

    master_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(suffix=master_path.suffix, dir=str(master_path.parent))
    os.close(fd)
    try:
        wb.save(temp_name)
        os.replace(temp_name, master_path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)

    audit_path.parent.mkdir(parents=True, exist_ok=True)
    with audit_path.open("a", encoding="utf-8") as audit:
        audit.write(json.dumps({"timestamp": datetime.now().isoformat(timespec="seconds"), "changes": changed_rows}) + "\n")
    return {"added": additions, "updated": updates, "changes": changed_rows}


def install_master(uploaded: bytes, master_path: Path, backup_dir: Path) -> None:
    if master_path.exists():
        backup_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(master_path, backup_dir / f"master_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    fd, temp_name = tempfile.mkstemp(suffix=".xlsx", dir=str(master_path.parent))
    os.close(fd)
    try:
        Path(temp_name).write_bytes(uploaded)
        wb = load_workbook(temp_name, data_only=False)
        _sheet_and_headers(wb)
        wb.close()
        os.replace(temp_name, master_path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)
